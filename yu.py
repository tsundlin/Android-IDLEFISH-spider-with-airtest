__author__ = "tsundlin"

# Using Python 3.7!!!!!!!!!!!!!
# There is bug for Python 3.9

import openpyxl
import sys
import csv

from airtest.cli.parser import cli_setup
from airtest.core.api import *
from poco.drivers.android.uiautomation import AndroidUiautomationPoco

# 爬取数量
HowManyItemsToLog = 10
# 搜索名字
SearchItemName = "微星GS65"
# 最低最高价格筛选
LowPrice = "5000"
HighPrice = "9000"
# 软件名字
AppName = "闲鱼"


def connectDevice():
    """
    connect device (夜神模拟器)
    """
    if not cli_setup():
        # 如果要换模拟器修改127.0.0.1:62001 到该模拟器指定的地址
        auto_setup(__file__, logdir=True, devices=[
            "Android://127.0.0.1:5037/127.0.0.1:62001?cap_method=JAVACAP^&^&ori_method=ADBORI",
        ])


def initializePoco():
    """
    initialize the Poco
    :return:
        AndroidUiautomationPoco : the poco
    """
    # poco安卓初始化
    poco = AndroidUiautomationPoco(
        use_airtest_input=True, screenshot_each_action=False)
    auto_setup(__file__)
    return poco


def openApp(poco):
    """
    open the app
    :param poco: the poco that we are using
    """
    poco(AppName).click()


def createExcel():
    """
    crete the result excel

    :return
        tuple(Workbook.active, Workbook) : return the excel that we are created
    """
    # 新建表格
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.cell(1, 1, "Title")
    worksheet.cell(1, 2, "Price")
    worksheet.cell(1, 3, "Location")
    return worksheet, workbook


def searchInApp(poco):
    """
    search the good with the conditions
    :param poco: the poco that we are using
    """
    poco("com.taobao.idlefish:id/bar_marquee_tx").click()
    text(SearchItemName)
    poco(text="搜索").click()
    poco(text="筛选").click()
    touch(Template(r"tpl1608125152364.png", record_pos=(-0.172, -0.359), resolution=(900, 1600)))

    text(LowPrice)
    touch(Template(r"tpl1608125165651.png", record_pos=(0.228, -0.364), resolution=(900, 1600)))

    text(HighPrice)
    poco(text="确定").click()


def collectData(work, poco):
    """
    collect the data the we want and put them into excel
    :param work: the excel thing we want to use
    :param poco: the poco that we are using
    """
    worksheet, workbook = work
    # excel 起始行数
    rowNum = 2

    while worksheet.max_row < HowManyItemsToLog:
        # 在出现的views 里找到我们需要的数据
        items = poco("android.widget.FrameLayout").offspring(type="android.view.View")
        res = []
        for item in items:
            text1 = item.get_text()
            if text1 is not None:
                res.append(text1)
        res = list(set(res))

        # 将数据分类
        for itemText in res:
            try:
                data = itemText.split('\n')
                itemName, price, location = data[0], data[1] + data[2], data[-1]
            except IndexError:
                continue
            # 写入excel
            worksheet.cell(rowNum, 1, itemName)
            worksheet.cell(rowNum, 2, price)
            worksheet.cell(rowNum, 3, location)
            workbook.save(SearchItemName + '.xlsx')
            rowNum = worksheet.max_row + 1
        res.clear()
        swipeScreen()


def swipeScreen():
    """
    swipe the screen that we can find more
    """
    swipe(v1=[450, 1586], v2=[450, 105], duration=3)
    sleep(2.0)


def main():
    """
    the main function
    """
    connectDevice()
    poco = initializePoco()
    openApp(poco)
    searchInApp(poco)
    work = createExcel()
    collectData(work, poco)


if __name__ == "__main__":
    main()
